#!/usr/bin/env python3
"""
BIM Data Execution - Bulk Company Lead Generation with Hierarchy (v4)

For each company:
  1. SEARCH wide (free) -> candidates with titles.
  2. RANK every candidate into hierarchy tiers.
  3. RESEARCH the top N distinct people (paid) -> emails + phones.
Output: ONE ROW PER CONTACT, grouped by company, ordered by tier, with Tier,
Hierarchy Role, name, title, email, phone, LinkedIn.

Speed: companies are processed in parallel by the web app (app.py).

REQUIREMENTS:  pip install requests openpyxl
"""

import argparse, csv, os, re, sys, time, json, sqlite3, threading, requests
from datetime import datetime, timezone
from openpyxl import load_workbook, Workbook

# ============================================================================
# MASTER STORE (SQLite cache)  - avoids paying to re-research known contacts.
# Key = company name + person name + hierarchy role (all lowercased).
# ============================================================================
MASTER_DB = os.environ.get("MASTER_DB_PATH") or os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "bim_master.db")
_DB_LOCK = threading.Lock()

def _norm(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower())

# Corporate/legal suffixes and connector words to strip when comparing company names.
_COMPANY_STRIP = re.compile(
    r"\b(inc|inc\.|incorporated|llc|l\.l\.c|ltd|limited|co|co\.|company|corp|corp\.|corporation|"
    r"plc|lp|llp|pllc|pc|gmbh|holdings|group|companies|services|solutions|the|and|&|of)\b",
    re.I)
_COMPANY_NOISE = re.compile(r"[^\w\s]+")

def _norm_company(s):
    """Aggressively normalise a company name for matching: lowercase, strip corporate
    suffixes and punctuation, collapse whitespace."""
    if not s: return ""
    x = s.lower()
    x = _COMPANY_NOISE.sub(" ", x)
    x = _COMPANY_STRIP.sub(" ", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x

def _norm_domain(s):
    """Strip protocol/www/path and lowercase, so any of these normalize to the same domain:
       https://www.example.com/about, example.com/, EXAMPLE.com -> example.com"""
    if not s: return ""
    x = s.strip().lower()
    x = re.sub(r"^https?://", "", x)
    x = re.sub(r"^www\.", "", x)
    x = x.split("/")[0].split("?")[0].strip()
    return x

def _domain_matches(asked, found):
    """Compare two domains loosely: exact match, or one ends with the other so that a
    subdomain still matches (e.g. 'careers.acme.com' matches 'acme.com')."""
    a = _norm_domain(asked); b = _norm_domain(found)
    if not a or not b: return False
    if a == b: return True
    return a.endswith("." + b) or b.endswith("." + a)

def _company_matches(asked, found, strict=True):
    """Smart, not strict: is the Seamless-reported company 'found' plausibly the same
    as the company we asked for? Returns True for exact match, substring match either
    way, or a strong token overlap. Blank 'found' -> False.

    When strict=False (used when a matching domain has already vouched for the
    candidate), we relax to any single distinctive shared token or substring."""
    if not found or not asked:
        return False
    a = _norm_company(asked); b = _norm_company(found)
    if not a or not b:
        return False
    if a == b or a in b or b in a:
        return True
    # Token overlap. Strip generic industry words that are shared across many firms.
    stop = {"construction", "contractors", "contracting", "builders", "building",
            "services", "consulting", "engineering", "engineers", "systems",
            "industries", "international", "national", "global", "associates",
            "partners", "enterprises", "management", "development", "properties",
            "realty", "real", "estate", "financial", "capital", "holdings",
            "group", "companies", "solutions"}
    def toks(x):
        return [t for t in x.split() if len(t) >= 3 and t not in stop]
    at, bt = toks(a), toks(b)
    if not at or not bt:
        # If stripping stop-words empties both sides, fall back to raw normalized
        # comparison (handles "The Group" vs "The Group Inc" style cases).
        raw_a = re.sub(r"\s+", " ", a).strip()
        raw_b = re.sub(r"\s+", " ", b).strip()
        return raw_a and raw_b and (raw_a == raw_b or raw_a in raw_b or raw_b in raw_a)
    shared = set(at) & set(bt)
    if not shared:
        return False
    # Strong: at least 2 shared distinctive tokens
    if len(shared) >= 2:
        return True
    # Relaxed mode: domain already matched, so any shared distinctive token is enough
    if not strict:
        return True
    # Strict mode: single shared token only OK if it's a long distinctive one
    if len(at) == 1 and len(bt) == 1:
        shared_tok = next(iter(shared))
        if len(shared_tok) >= 5:
            return True
    return False

def _cache_key(company, first, last, role):
    return f"{_norm(company)}|{_norm(first)} {_norm(last)}|{_norm(role)}"

def db_connect():
    conn = sqlite3.connect(MASTER_DB, timeout=30, check_same_thread=False)
    conn.execute("""CREATE TABLE IF NOT EXISTS contacts (
        cache_key TEXT PRIMARY KEY,
        company TEXT, tier TEXT, role TEXT,
        first_name TEXT, last_name TEXT, job_title TEXT,
        email TEXT, phone TEXT, linkedin TEXT, domain TEXT,
        saved_at TEXT )""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_company ON contacts(company)")
    return conn

def cache_lookup_company(company, lock_timeout=3):
    """Return list of saved contact rows for a company (empty if none).
    If the DB lock cannot be acquired in `lock_timeout` seconds, we treat this
    as a cache miss so a stuck DB never freezes a worker."""
    got_lock = _DB_LOCK.acquire(timeout=lock_timeout)
    if not got_lock:
        log(f"      [cache] {company!r}: DB lock timeout - treating as cache miss")
        return []
    try:
        conn = db_connect()
        try:
            cur = conn.execute("SELECT company,tier,role,first_name,last_name,job_title,"
                               "email,phone,linkedin,domain,saved_at FROM contacts "
                               "WHERE company=?", (_norm(company),))
            out = []
            for r in cur.fetchall():
                out.append({"Company Name": company, "Tier": r[1], "Hierarchy Role": r[2],
                            "First Name": r[3], "Last Name": r[4], "Job Title": r[5],
                            "Email": r[6], "Phone Number": r[7], "LinkedIn Profile": r[8],
                            "Company Domain": r[9],
                            "Source": f"From cache (saved {(r[10] or '')[:10]})", "Note": ""})
            return out
        finally:
            conn.close()
    finally:
        _DB_LOCK.release()

def cache_save(rows, lock_timeout=5):
    """Upsert freshly-researched contact rows into the master store.
    Skips silently if the DB lock is unavailable - caller shouldn't be blocked."""
    now = datetime.now(timezone.utc).isoformat()
    got_lock = _DB_LOCK.acquire(timeout=lock_timeout)
    if not got_lock:
        log(f"      [cache] save skipped - DB lock timeout")
        return
    try:
        conn = db_connect()
        try:
            for r in rows:
                if not (r.get("First Name") or r.get("Last Name")):
                    continue  # skip note-only / no-match rows
                key = _cache_key(r["Company Name"], r.get("First Name"),
                                 r.get("Last Name"), r.get("Hierarchy Role"))
                conn.execute("""INSERT INTO contacts
                    (cache_key,company,tier,role,first_name,last_name,job_title,
                     email,phone,linkedin,domain,saved_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(cache_key) DO UPDATE SET
                     email=excluded.email, phone=excluded.phone, linkedin=excluded.linkedin,
                     domain=excluded.domain, job_title=excluded.job_title,
                     tier=excluded.tier, saved_at=excluded.saved_at""",
                    (key, _norm(r["Company Name"]), str(r.get("Tier","")), r.get("Hierarchy Role",""),
                     r.get("First Name",""), r.get("Last Name",""), r.get("Job Title",""),
                     r.get("Email",""), r.get("Phone Number",""), r.get("LinkedIn Profile",""),
                     r.get("Company Domain",""), now))
            conn.commit()
        finally:
            conn.close()
    finally:
        _DB_LOCK.release()

def cache_stats():
    with _DB_LOCK:
        conn = db_connect()
        try:
            n = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
            c = conn.execute("SELECT COUNT(DISTINCT company) FROM contacts").fetchone()[0]
            # Also report how many are "phoneless" so we can see cache health
            bad = conn.execute("SELECT COUNT(*) FROM contacts WHERE "
                               "(email IS NULL OR email='') AND "
                               "(phone IS NULL OR phone='')").fetchone()[0]
            return {"contacts": n, "companies": c, "phoneless": bad}
        finally:
            conn.close()

def cache_purge_phoneless():
    """Delete all cached contacts that have neither phone nor email.
    Returns (deleted_count, remaining_count).
    Companies with ONLY phoneless rows will be re-researched on next run."""
    with _DB_LOCK:
        conn = db_connect()
        try:
            cur = conn.execute("DELETE FROM contacts WHERE "
                               "(email IS NULL OR email='') AND "
                               "(phone IS NULL OR phone='')")
            deleted = cur.rowcount
            conn.commit()
            remaining = conn.execute("SELECT COUNT(*) FROM contacts").fetchone()[0]
            return deleted, remaining
        finally:
            conn.close()

def export_master_xlsx(path):
    with _DB_LOCK:
        conn = db_connect()
        try:
            cur = conn.execute("SELECT company,tier,role,first_name,last_name,job_title,"
                               "email,phone,linkedin,domain,saved_at FROM contacts "
                               "ORDER BY company,tier")
            wb = Workbook(); ws = wb.active; ws.title = "Master"
            ws.append(["Company","Tier","Hierarchy Role","First Name","Last Name","Job Title",
                       "Email","Phone Number","LinkedIn Profile","Company Domain","Saved At"])
            for r in cur.fetchall():
                ws.append(list(r))
            wb.save(path)
        finally:
            conn.close()

# ============================================================================
# CONFIG  (Seamless auth: header named "token", raw key)
# ============================================================================
SEAMLESS_BASE = "https://api.seamless.ai/api/client/v1"
EP_SEARCH   = f"{SEAMLESS_BASE}/search/contacts"
EP_RESEARCH = f"{SEAMLESS_BASE}/contacts/research"
EP_POLL     = f"{SEAMLESS_BASE}/contacts/research/poll"

AUTH_HEADER_NAME  = "token"
AUTH_HEADER_VALUE = "{key}"

SEARCH_LIMIT              = 25        # candidates to pull & rank per company
MAX_CONTACTS_PER_COMPANY  = 5         # how many people to RESEARCH per company (credit driver)
SKIP_DEDUP                = False
HTTP_TIMEOUT              = int(os.environ.get('HTTP_TIMEOUT', '15'))   # per HTTP attempt - kept low so hangs fail fast
HTTP_RETRIES              = int(os.environ.get('HTTP_RETRIES', '1'))    # so max wait per call = 20+2+20 = 42s
POLL_INTERVAL_SECONDS     = int(os.environ.get("POLL_INTERVAL_SECONDS", "8"))
POLL_MAX_ATTEMPTS         = int(os.environ.get("POLL_MAX_ATTEMPTS", "20"))
DELAY_BETWEEN_COMPANIES   = 0
DONE_STATUSES             = {"done", "error", "missing", "duplicate"}
USABLE_STATUSES           = {"done", "duplicate"}

COMPANY_COLUMN_CANDIDATES = ["company", "company name", "companyname", "firm", "firm name"]
DOMAIN_COLUMN_CANDIDATES = ["website", "domain", "url", "company website", "company domain",
                            "web", "site"]
EMAIL_COLUMN_CANDIDATES = ["email", "email address", "contact email", "e-mail"]

def _domain_from_email(s):
    """Extract the domain portion of an email, so 'john@bcsconstructiongroup.com' -> 'bcsconstructiongroup.com'.
    Returns '' if the string isn't email-shaped."""
    if not s: return ""
    s = str(s).strip()
    if "@" not in s: return ""
    part = s.split("@", 1)[1].strip()
    return _norm_domain(part)

# ============================================================================
# HIERARCHY  (user-defined, priority order: index 1 = highest priority)
#   ONLY titles in this list match. Anything not listed is not selected.
#   Ordered patterns; the FIRST pattern a title matches wins its tier.
#   More specific titles are listed before broader ones so they win correctly.
# ============================================================================
HIERARCHY_TITLES = [
    # --- BIM / VDC / CAD ---
    ("BIM Director",                    r"bim director|director of bim"),
    ("VDC Director",                    r"vdc director|director of vdc"),
    ("Virtual Design Manager",          r"virtual design (and construction )?manager"),
    ("VDC Manager",                     r"vdc manager"),
    ("BIM Manager",                     r"bim manager"),
    ("CAD Manager",                     r"cad manager"),
    ("BIM Coordinator",                 r"bim coordinator"),
    ("CAD Coordinator",                 r"cad coordinator"),
    # --- Pre-Construction ---
    ("VP of Pre-Construction",          r"(vice president|vp).*(pre-?con(struction)?)|(pre-?con(struction)?).*(vice president|vp)"),
    ("Director of Pre-Construction",    r"(director of pre-?con(struction)?)|(pre-?con(struction)?\s*director)"),
    ("Pre-Construction Executive",      r"pre-?con(struction)?\s*executive"),
    ("Pre-Construction Lead",           r"pre-?con(struction)?\s*lead"),
    ("Pre-Construction Operations Mgr", r"pre-?con(struction)?\s*operations\s*manager"),
    ("Pre-Construction Manager",        r"pre-?con(struction)?\s*manager"),
    ("Pre-Construction Coordinator",    r"pre-?con(struction)?\s*coordinator"),
    # --- Division / General Management ---
    ("VP - Division",                   r"(vice president|vp)\s*[-\u2013]?\s*division"),
    ("VP - Region",                     r"(vice president|vp)\s*[-\u2013]?\s*region"),
    ("Regional Division Manager",       r"regional division manager"),
    ("Mechanical Division Manager",     r"mechanical division manager"),
    ("Electrical Division Manager",     r"electrical division manager"),
    ("Plumbing Division Manager",       r"plumbing division manager"),
    ("Division Director",               r"division director"),
    ("Division Manager",                r"division manager"),
    ("Vertical Manager",                r"vertical manager"),
    ("General Manager",                 r"general manager"),
    # --- Estimating ---
    ("Chief Estimating Officer",        r"chief estimating officer"),
    ("Chief Estimator",                 r"chief estimator"),
    ("Director of Estimating",          r"(director of estimating)|(estimating director)"),
    ("Estimating Manager",              r"estimating manager"),
    ("Senior Estimator",               r"(senior|sr\.?) estimator"),
    ("Lead Estimator",                  r"lead estimator"),
    ("Mechanical Estimator",            r"mechanical estimator"),
    ("HVAC Estimator",                  r"hvac estimator"),
    ("Electrical Estimator",            r"electrical estimator"),
    ("Plumbing Estimator",              r"plumbing estimator"),
    ("Estimator",                       r"\bestimator\b"),
    # --- Project Management ---
    ("Director of Project Management",  r"director of project management"),
    ("Project Executive",               r"project executive"),
    ("Senior Project Manager",         r"(senior|sr\.?) project manager"),
    ("Project Manager",                 r"\bproject manager\b"),
    ("Assistant Project Manager",       r"assistant project manager"),
    ("Senior Construction Manager",     r"(senior|sr\.?) construction manager"),
    ("Construction Manager",            r"construction manager"),
    ("Senior Project Engineer",         r"(senior|sr\.?) project engineer"),
    ("Project Engineer",                r"project engineer"),
    ("Project Coordinator",             r"project coordinator"),
    # --- Engineering ---
    ("Director of Engineering",         r"director of engineering"),
    ("Chief Engineer",                  r"chief engineer"),
    ("Engineering Manager",             r"engineering manager"),
    ("Senior Engineer",                 r"(senior|sr\.?) engineer"),
    ("Mechanical Engineer",             r"mechanical engineer"),
    ("Electrical Engineer",             r"electrical engineer"),
    ("Plumbing Engineer",               r"plumbing engineer"),
    ("Design Engineer",                 r"design engineer"),
    # --- Operations ---
    ("VP of Operations",                r"(vice president|vp).*operations|operations.*(vice president|vp)"),
    ("Chief Operating Officer",         r"chief operating officer|\bcoo\b"),
    ("Director of Operations",          r"(director of operations)|(operations director)"),
    ("Senior Operations Manager",       r"(senior|sr\.?) operations manager"),
    ("Operations Manager",              r"operations manager"),
    # --- Executive / President ---
    ("Executive Vice President",        r"executive vice president|\bevp\b"),
    ("Senior Vice President",           r"senior vice president|\bsvp\b"),
    ("Regional President",              r"regional president"),
    ("Company President",               r"company president"),
    ("President",                       r"\bpresident\b"),
    ("Vice President",                  r"vice president|\bvp\b"),
    # --- Owner / Founder / CEO (+ assistants to them, listed before CEO/Owner so they win) ---
    ("Executive Assistant to CEO",      r"executive assistant to (the )?ceo"),
    ("Assistant to CEO",                r"assistant to (the )?ceo"),
    ("Assistant to Owner",              r"assistant to (the )?owner"),
    ("Chief Executive Officer",         r"chief executive officer|\bceo\b"),
    ("Owner",                           r"\bowner\b"),
    ("Founder",                         r"\bfounder\b"),
]
# Compile into tiered rules (tier = position in the list, 1-based)
HIERARCHY = [(i + 1, label, re.compile(pat, re.I))
             for i, (label, pat) in enumerate(HIERARCHY_TITLES)]
TIER_LABEL = {i + 1: label for i, (label, _) in enumerate(HIERARCHY_TITLES)}
LAST_RESORT_TIER = 10_000   # unused now (kept so other code references don't break)
BLACKLIST = []              # no exclusions - only listed titles match
LAST_RESORT = []

OUTPUT_COLUMNS = ["Company Name", "Tier", "Hierarchy Role", "First Name", "Last Name",
                  "Job Title", "Email", "Phone Number", "LinkedIn Profile",
                  "Company Domain", "Source", "Note"]


def log(msg): print(msg, flush=True)

def auth_headers(api_key):
    return {AUTH_HEADER_NAME: AUTH_HEADER_VALUE.format(key=api_key), "Content-Type": "application/json"}

def load_companies(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True); ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    elif ext in (".csv", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = [r for r in csv.reader(f, delimiter=delim)]
    else:
        raise ValueError(f"Unsupported file type: {ext} (use .xlsx or .csv)")
    if not rows: raise ValueError("File is empty.")
    header = [("" if c is None else str(c)).strip() for c in rows[0]]
    col_idx = None
    dom_idx = None
    email_idx = None
    for i, h in enumerate(header):
        hl = h.strip().lower()
        if col_idx is None and hl in COMPANY_COLUMN_CANDIDATES: col_idx = i
        if dom_idx is None and hl in DOMAIN_COLUMN_CANDIDATES: dom_idx = i
        if email_idx is None and hl in EMAIL_COLUMN_CANDIDATES: email_idx = i
    if col_idx is None:
        raise ValueError(f'No company column found. Headers: {header}. Expected one of: {COMPANY_COLUMN_CANDIDATES}')
    companies = []
    seen = set()
    for r in rows[1:]:
        if col_idx < len(r) and r[col_idx] is not None:
            name = str(r[col_idx]).strip()
            # Skip the template hint row created by _universal_ensure_input
            if name.startswith("(") and name.endswith(")"):
                continue
            domain = ""
            email_cell = ""
            web_cell = ""
            if email_idx is not None and email_idx < len(r) and r[email_idx] is not None:
                email_cell = str(r[email_idx]).strip()
            if dom_idx is not None and dom_idx < len(r) and r[dom_idx] is not None:
                web_cell = str(r[dom_idx]).strip()
            # Rule: pick Email OR Website, not both. Prefer Email if both filled.
            if email_cell and web_cell:
                log(f"      [input] '{name}': both Email and Website provided - using Email, ignoring Website.")
                domain = _domain_from_email(email_cell)
            elif email_cell:
                domain = _domain_from_email(email_cell)
                if not domain:
                    log(f"      [input] '{name}': Email value {email_cell!r} isn't a valid email; ignoring.")
            elif web_cell:
                domain = _norm_domain(web_cell)
            key = (name.lower(), domain)
            if name and key not in seen:
                seen.add(key); companies.append((name, domain))
    if not companies: raise ValueError("Company column found, but no non-empty company names.")
    # Enforce the 1-200 company cap for this tool
    MAX_INPUT_COMPANIES = 200
    if len(companies) > MAX_INPUT_COMPANIES:
        raise ValueError(f"Too many companies: file has {len(companies)}, max allowed is {MAX_INPUT_COMPANIES}. "
                         f"Split your file into batches of {MAX_INPUT_COMPANIES} or fewer.")
    return companies


# ---------------------------------------------------------------------------
# Ranking (free, on search titles)
# ---------------------------------------------------------------------------
def _match_text(contact):
    return f"{contact.get('title','') or ''} {contact.get('department','') or ''}".strip()

def tier_of(contact):
    """Return the hierarchy tier for a contact's title, or None if the title isn't
    in the defined hierarchy. First matching pattern (highest priority) wins."""
    title = (contact.get("title") or "").strip()
    if not title:
        return None
    for tier, label, pat in HIERARCHY:
        if pat.search(title):
            return tier
    return None

def rank_candidates(contacts):
    """Return list of (tier, contact) for all contacts whose title is in the hierarchy,
    best tier first, de-duplicated by person."""
    ranked = []
    seen = set()
    for c in contacts:
        t = tier_of(c)
        if t is None:
            continue
        key = (c.get("name") or "").strip().lower() or (c.get("liUrl") or "")
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        ranked.append((t, c))
    ranked.sort(key=lambda x: x[0])
    return ranked


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------
def _split_name(full):
    parts = (full or "").split()
    if not parts: return "", ""
    if len(parts) == 1: return parts[0], ""
    return parts[0], parts[-1]

_EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_PHONE_RE = re.compile(r"(?:\+?\d[\d\-\.\s\(\)]{6,}\d)")

def _flat_str(v):
    """Pull a usable string out of a value that may be a str, list, or dict."""
    if v is None: return ""
    if isinstance(v, str): return v.strip()
    if isinstance(v, (int, float)): return str(v)
    if isinstance(v, list):
        for item in v:
            s = _flat_str(item)
            if s: return s
        return ""
    if isinstance(v, dict):
        for item in v.values():
            s = _flat_str(item)
            if s: return s
        return ""
    return str(v).strip()

def _find_email(*objs):
    for o in objs:
        if not isinstance(o, dict): continue
        for k in ("email", "email1", "emailAddress", "workEmail", "primaryEmail", "email2", "emails"):
            s = _flat_str(o.get(k))
            if s and _EMAIL_RE.search(s): return _EMAIL_RE.search(s).group(0)
    for o in objs:
        if not isinstance(o, dict): continue
        for k, v in o.items():
            if "email" in k.lower():
                s = _flat_str(v)
                if s and _EMAIL_RE.search(s): return _EMAIL_RE.search(s).group(0)
    for o in objs:
        if not isinstance(o, dict): continue
        for v in o.values():
            s = _flat_str(v)
            if s and _EMAIL_RE.search(s): return _EMAIL_RE.search(s).group(0)
    return ""

def _find_phone(*objs):
    for o in objs:
        if not isinstance(o, dict): continue
        for k in ("contactPhone1", "contactPhone2", "contactPhone3", "phone", "phoneNumber",
                  "mobilePhone", "directPhone", "workPhone", "companyPhone", "phones"):
            s = _flat_str(o.get(k))
            if s: return s
    for o in objs:
        if not isinstance(o, dict): continue
        for k, v in o.items():
            if "phone" in k.lower():
                s = _flat_str(v)
                if s: return s
    return ""

def _find_linkedin(*objs):
    for o in objs:
        if not isinstance(o, dict): continue
        for k in ("lIProfileUrl", "liUrl", "linkedin", "linkedInUrl", "linkedinUrl"):
            s = _flat_str(o.get(k))
            if s and "linkedin" in s.lower(): return s
    for o in objs:
        if not isinstance(o, dict): continue
        for v in o.values():
            s = _flat_str(v)
            if s and "linkedin.com/in/" in s.lower(): return s
    return ""

def contact_row(company, tier, search_c, enriched, note=""):
    enriched = enriched or {}
    fn = enriched.get("firstName") or search_c.get("firstName")
    ln = enriched.get("lastName") or search_c.get("lastName")
    if not fn and not ln:
        fn, ln = _split_name(search_c.get("name") or enriched.get("name"))
    # scan BOTH the enriched (research) and search objects, any field name
    email = _find_email(enriched, search_c)
    phone = _find_phone(enriched, search_c)
    linkedin = _find_linkedin(enriched, search_c)
    return {
        "Company Name": company,
        "Tier": tier,
        "Hierarchy Role": TIER_LABEL.get(tier, ""),
        "First Name": fn or "", "Last Name": ln or "",
        "Job Title": enriched.get("title") or search_c.get("title") or "",
        "Email": email, "Phone Number": phone,
        "LinkedIn Profile": linkedin,
        "Company Domain": enriched.get("companyDomain") or search_c.get("domain") or "",
        "Source": "Researched now",
        "Note": note or ("" if email else "No email returned by research."),
    }

def note_row(company, note):
    row = {k: "" for k in OUTPUT_COLUMNS}
    row["Company Name"] = company; row["Note"] = note
    return row


# ---------------------------------------------------------------------------
# Seamless API (with retry on timeout)
# ---------------------------------------------------------------------------
def _request(method, url, headers, **kw):
    """HTTP call with timeout + retries + logging of every attempt."""
    last = None
    total_attempts = HTTP_RETRIES + 1
    for attempt in range(1, total_attempts + 1):
        _t0 = time.time()
        log(f"      [http] {method} {url.split('/')[-1]} attempt {attempt}/{total_attempts} ...")
        try:
            r = requests.request(method, url, headers=headers, timeout=HTTP_TIMEOUT, **kw)
            _elapsed = time.time() - _t0
            log(f"      [http] {method} {url.split('/')[-1]} attempt {attempt}: HTTP {r.status_code} in {_elapsed:.1f}s")
            if r.status_code >= 500:
                last = RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
                if attempt < total_attempts:
                    time.sleep(2); continue
                break
            return r
        except (requests.Timeout, requests.ConnectionError) as e:
            _elapsed = time.time() - _t0
            log(f"      [http] {method} {url.split('/')[-1]} attempt {attempt}: FAILED after {_elapsed:.1f}s: {type(e).__name__}: {e}")
            last = e
            if attempt < total_attempts:
                time.sleep(2)
    raise last

def safe_json(resp, label):
    if resp.status_code >= 400:
        raise RuntimeError(f"{label} HTTP {resp.status_code}: {resp.text[:300]}")
    try:
        return resp.json()
    except Exception:
        raise RuntimeError(f"{label}: non-JSON response: {resp.text[:300]}")

def search_candidates(company, headers, limit=None, domain=""):
    limit = limit or SEARCH_LIMIT
    body = {"companyName": company, "companyDomain": _norm_domain(domain) if domain else "",
            "limit": limit}
    log(f"      [search] {company!r} (domain={domain!r}) - calling Seamless...")
    try:
        data = safe_json(_request("POST", EP_SEARCH, headers, json=body), "search")
        cands = data.get("data") or data.get("contacts") or data.get("results") or []
        log(f"      [search] {company!r} - got {len(cands)} candidate(s).")
        return cands
    except Exception as e:
        log(f"      [search] {company!r} - FAILED: {e}")
        raise

def research_ids(search_result_ids, headers):
    body = {"searchResultIds": list(search_result_ids), "skipDeduplicationCheck": SKIP_DEDUP}
    log(f"      [research] submitting {len(search_result_ids)} id(s) to Seamless...")
    try:
        data = safe_json(_request("POST", EP_RESEARCH, headers, json=body), "research-submit")
    except Exception as e:
        log(f"      [research] submit FAILED: {e}")
        raise
    request_ids = data.get("requestIds") or (data.get("data") or {}).get("requestIds") or []
    if not request_ids:
        err = data.get("message") or data.get("error") or json.dumps(data)[:200]
        log(f"      [research] no requestIds returned. Raw: {err}")
        raise RuntimeError(f"research submit returned no requestIds. Raw: {err}")
    log(f"      [research] submitted OK, requestIds={request_ids}")
    return request_ids

def poll_research(request_ids, headers, interval=None, attempts=None):
    interval = interval or POLL_INTERVAL_SECONDS
    attempts = attempts or POLL_MAX_ATTEMPTS
    params = {"requestIds": ",".join(str(x) for x in request_ids)}
    results = []
    deadline = time.time() + (attempts * interval) + 30   # hard wall-clock limit
    for attempt in range(1, attempts + 1):
        if time.time() > deadline:
            log(f"      [poll] hard deadline reached at attempt {attempt} - skipping company.")
            break
        try:
            data = safe_json(_request("GET", EP_POLL, headers, params=params), "poll")
        except Exception as e:
            log(f"      [poll] attempt {attempt}/{attempts}: request error: {e}")
            time.sleep(interval)
            continue
        results = data.get("data") or []
        statuses = [r.get("status") for r in results] if isinstance(results, list) else []
        log(f"      [poll] attempt {attempt}/{attempts}: statuses={statuses}")
        if isinstance(results, list) and results and all(s in DONE_STATUSES for s in statuses):
            log(f"      [poll] complete after {attempt} attempt(s).")
            return results
        time.sleep(interval)
    log(f"      [poll] gave up after {attempts} attempts - using partial data.")
    return results if isinstance(results, list) else []


def _id_of(c):
    return (c.get("searchResultId") or c.get("searchResultID")
            or c.get("id") or c.get("contactId"))

def process_company(company, headers, search_limit=None, poll_interval=None,
                    poll_attempts=None, min_rank=None, max_contacts=None,
                    preview=False, use_cache=True, company_domain=""):
    """Return (rows, kind). rows is a list (one per contact) or a single note row.
    kind: nocontacts | nomatch | would | found | cached | error(handled by caller)

    If company_domain is given, it's passed to Seamless search AND each candidate must
    match on BOTH company name and domain (safest 'use website + name together' mode)."""
    log(f"      [process] {company!r} - entering process_company")
    max_contacts = max_contacts or MAX_CONTACTS_PER_COMPANY
    company_domain = _norm_domain(company_domain) if company_domain else ""

    # 0) MASTER STORE: if we already researched this company, reuse it for free.
    # But only if the cached rows actually contain useful contact data (phone or email).
    # Otherwise the cache was populated when only search-tier data was available,
    # and we should re-research to get real phones/emails.
    if use_cache and not preview:
        cached = cache_lookup_company(company)
        if cached:
            # Check if cached rows have at least one row with phone or email
            has_contact_info = any(
                (r.get("Email", "").strip() or r.get("Phone Number", "").strip())
                for r in cached
            )
            if has_contact_info:
                log(f"      [cache] {company}: returning {len(cached)} cached contact(s) (free)")
                return cached, "cached"
            else:
                log(f"      [cache] {company}: cache has {len(cached)} row(s) but no phones/emails - re-researching")
                # fall through to fresh research

    candidates = search_candidates(company, headers, limit=search_limit, domain=company_domain)
    if not candidates:
        return [note_row(company, "No contacts found for this company in Seamless search.")], "nocontacts"

    # --- COMPANY-MATCH FILTER: drop candidates whose Seamless "company" field
    # doesn't plausibly match the company we asked for. Fixes "wrong person for
    # the company" (e.g. searching BCS Construction Group and getting BCS Financial).
    #
    # Matching strategy (in order of trust):
    #   1. If a domain was supplied AND the candidate's domain matches -> KEEP.
    #      Domain is the strongest possible signal; relaxed name check as sanity.
    #   2. If a domain was supplied but candidate has no domain -> fall back to
    #      strict name match (Seamless often omits domain even for real hits).
    #   3. No domain supplied -> use strict name match only.
    kept, dropped = [], []
    for c in candidates:
        cand_co = c.get("companyName") or c.get("company") or c.get("companyOriginal") or ""
        cand_dom = c.get("companyDomain") or c.get("domain") or ""
        keep = False
        reason = ""
        if company_domain:
            if cand_dom and _domain_matches(company_domain, cand_dom):
                # Domain match: relaxed name check as sanity (catches Seamless mixups)
                if _company_matches(company, cand_co, strict=False) or not cand_co:
                    keep = True
                else:
                    reason = "domain-ok-but-name-mismatch"
            elif cand_dom:
                # Candidate has a domain and it doesn't match ours -> definitely wrong
                reason = "wrong-domain"
            else:
                # No candidate domain: fall back to STRICT name match
                if _company_matches(company, cand_co, strict=True):
                    keep = True
                else:
                    reason = "no-domain-and-name-mismatch"
        else:
            # No domain supplied at all: strict name match only
            if _company_matches(company, cand_co, strict=True):
                keep = True
            else:
                reason = "name-mismatch"
        if keep:
            kept.append(c)
        else:
            dropped.append((c.get("name") or "?", cand_co, cand_dom, reason))
    if dropped:
        # Log up to 3 examples of what got dropped, so the user can see the filter working
        examples = ", ".join(f"{n} @ {co!r} ({r})" for n, co, _d, r in dropped[:3])
        log(f"      [filter] {company}: dropped {len(dropped)} candidate(s) at other companies "
            f"(e.g. {examples})")
    candidates = kept
    if not candidates:
        msg = ("All Seamless candidates worked at other companies; skipped to avoid "
               "wrong-company contacts.")
        if company_domain:
            msg = (f"All Seamless candidates for '{company}' failed the name-or-domain "
                   f"match against {company_domain}. Skipped.")
        return [note_row(company, msg)], "nomatch"

    ranked = rank_candidates(candidates)
    if min_rank is not None:
        ranked = [(t, c) for t, c in ranked if t <= min_rank]
    if not ranked:
        return [note_row(company, "No contact matched the hierarchy (no one at this company "
                         "held a title in your hierarchy list).")], "nomatch"

    ranked = ranked[:max_contacts]

    if preview:
        rows = []
        # in preview, show whether each would be free (cached) or paid
        cached = {(_norm(r["First Name"]) + " " + _norm(r["Last Name"])).strip()
                  for r in (cache_lookup_company(company) if use_cache else [])}
        for t, c in ranked:
            name = (c.get("name") or f"{c.get('firstName','')} {c.get('lastName','')}").strip()
            fn, ln = _split_name(c.get("name")) if not c.get("firstName") else (c.get("firstName"), c.get("lastName"))
            is_cached = (_norm(fn) + " " + _norm(ln)).strip() in cached
            row = contact_row(company, t, c, {},
                    note=(f"PREVIEW: already saved - would reuse FREE."
                          if is_cached else f"PREVIEW: would research {name} ({TIER_LABEL.get(t)}). ~1 credit."))
            row["Source"] = "From cache (free)" if is_cached else "Would research"
            rows.append(row)
        return rows, "would"

    # research all selected in ONE batch call (fast) then poll once
    ranked_with_id = [(t, c) for t, c in ranked if _id_of(c)]
    ids = [_id_of(c) for _, c in ranked_with_id]
    enriched_by_key = {}
    ordered_results = []
    if ids:
        req = research_ids(ids, headers)
        results = poll_research(req, headers, interval=poll_interval, attempts=poll_attempts)
        # Diagnostic: report how many results came back with phones vs emails
        phone_count = 0; email_count = 0
        for r in results:
            if not isinstance(r, dict):
                continue
            c = r.get("contact") if isinstance(r.get("contact"), dict) else r
            if _find_phone(c): phone_count += 1
            if _find_email(c): email_count += 1
        log(f"      [research] {company}: got {len(results)} result(s), "
            f"{phone_count} with phone, {email_count} with email")
        for r in results:
            if not isinstance(r, dict):
                continue
            status = r.get("status")
            c = r.get("contact") if isinstance(r.get("contact"), dict) else r
            # Only skip explicit failures that carry no data.
            if status in ("error", "missing", "researching") and not (
                    c.get("email") or c.get("email1") or c.get("firstName")):
                continue
            ordered_results.append(c)
            # The poll item's TOP-LEVEL searchResultId matches the search id we submitted.
            top_id = r.get("searchResultId") or r.get("searchResultID")
            if top_id:
                enriched_by_key[str(top_id)] = c
            # also register any ids inside the contact object, and the name
            for k in (c.get("searchResultId"), c.get("id"), c.get("contactId")):
                if k:
                    enriched_by_key[str(k)] = c
            nm = _norm(f"{c.get('firstName','')} {c.get('lastName','')}") or _norm(c.get("name"))
            if nm.strip():
                enriched_by_key["name:" + nm] = c

    rows = []
    for idx, (t, c) in enumerate(ranked_with_id):
        cid = _id_of(c)
        nm = _norm(f"{c.get('firstName','')} {c.get('lastName','')}") or _norm(c.get("name"))
        enr = (enriched_by_key.get(str(cid))
               or enriched_by_key.get("name:" + nm)
               or (ordered_results[idx] if idx < len(ordered_results) else None)  # order fallback
               or {})
        rows.append(contact_row(company, t, c, enr))
    # include any ranked contacts that had no id (rare) so nothing is dropped
    for t, c in ranked:
        if _id_of(c) is None:
            rows.append(contact_row(company, t, c, {}))

    # SAVE newly researched contacts to the master store for next time
    if use_cache:
        try: cache_save(rows)
        except Exception: pass
    return rows, "found"


def write_xlsx(rows, path):
    wb = Workbook(); ws = wb.active; ws.title = "Contacts"; ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(k, "") for k in OUTPUT_COLUMNS])
    wb.save(path)


# ============================================================================
# GOOGLE DRIVE upload (optional). Enabled when these env vars are set:
#   GDRIVE_CLIENT_ID, GDRIVE_CLIENT_SECRET, GDRIVE_REFRESH_TOKEN, GDRIVE_FOLDER_ID
# Uploads a local file into the shared Drive folder. Uses only the refresh token
# (no browser needed on the server). Fails soft: never breaks a run.
# ============================================================================
GDRIVE_TOKEN_URI = "https://oauth2.googleapis.com/token"
GDRIVE_UPLOAD_URL = "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&supportsAllDrives=true"

def drive_enabled():
    return all(os.environ.get(k) for k in
               ("GDRIVE_CLIENT_ID", "GDRIVE_CLIENT_SECRET", "GDRIVE_REFRESH_TOKEN", "GDRIVE_FOLDER_ID"))

def _drive_access_token():
    r = requests.post(GDRIVE_TOKEN_URI, data={
        "client_id": os.environ["GDRIVE_CLIENT_ID"],
        "client_secret": os.environ["GDRIVE_CLIENT_SECRET"],
        "refresh_token": os.environ["GDRIVE_REFRESH_TOKEN"],
        "grant_type": "refresh_token",
    }, timeout=30)
    r.raise_for_status()
    return r.json()["access_token"]

def drive_upload(local_path, drive_name=None, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
    """Upload a file to the configured Drive folder. Returns the Drive file link, or None."""
    if not drive_enabled():
        return None
    try:
        token = _drive_access_token()
        meta = {"name": drive_name or os.path.basename(local_path),
                "parents": [os.environ["GDRIVE_FOLDER_ID"]]}
        with open(local_path, "rb") as f:
            data = f.read()
        boundary = "----bimboundary7hf83n"
        body = (
            (f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n"
             f"{json.dumps(meta)}\r\n").encode()
            + f"--{boundary}\r\nContent-Type: {mime}\r\n\r\n".encode()
            + data + f"\r\n--{boundary}--".encode()
        )
        r = requests.post(GDRIVE_UPLOAD_URL, headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": f"multipart/related; boundary={boundary}",
        }, data=body, timeout=120)
        r.raise_for_status()
        fid = r.json().get("id")
        return f"https://drive.google.com/file/d/{fid}/view" if fid else None
    except Exception as e:
        log(f"      [drive] upload failed: {e}")
        return None


# ---------------------------------------------------------------------------
# PERMANENT MASTER stored as a single Excel file in Drive (survives free-tier
# restarts). Flow each run: pull master from Drive -> merge into local SQLite ->
# push updated master back to the same Drive file.
# ---------------------------------------------------------------------------
MASTER_DRIVE_NAME = "BIM_MASTER_CONTACTS.xlsx"
GDRIVE_FILES_URL = "https://www.googleapis.com/drive/v3/files"
GDRIVE_UPDATE_URL = "https://www.googleapis.com/upload/drive/v3/files/{fid}?uploadType=media&supportsAllDrives=true"
_MASTER_SYNC_LOCK = threading.Lock()

def _drive_find_master(token):
    """Return the fileId of the master Excel in the folder, or None."""
    folder = os.environ["GDRIVE_FOLDER_ID"]
    q = f"name='{MASTER_DRIVE_NAME}' and '{folder}' in parents and trashed=false"
    r = requests.get(GDRIVE_FILES_URL, headers={"Authorization": f"Bearer {token}"},
                     params={"q": q, "fields": "files(id,name)",
                             "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"},
                     timeout=30)
    r.raise_for_status()
    files = r.json().get("files", [])
    return files[0]["id"] if files else None

def _drive_download(token, fid, dest_path):
    r = requests.get(f"{GDRIVE_FILES_URL}/{fid}",
                     headers={"Authorization": f"Bearer {token}"},
                     params={"alt": "media", "supportsAllDrives": "true"}, timeout=120)
    r.raise_for_status()
    with open(dest_path, "wb") as f:
        f.write(r.content)

def _import_master_xlsx_into_db(path):
    """Load a master Excel (exported by export_master_xlsx) back into the SQLite cache."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    hdr = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}
    def g(r, key):
        i = idx.get(key)
        return (r[i] if (i is not None and i < len(r) and r[i] is not None) else "")
    imported = []
    for r in rows[1:]:
        if not any(v not in (None, "") for v in r):
            continue
        imported.append({
            "Company Name": g(r, "Company"), "Tier": g(r, "Tier"),
            "Hierarchy Role": g(r, "Hierarchy Role"), "First Name": g(r, "First Name"),
            "Last Name": g(r, "Last Name"), "Job Title": g(r, "Job Title"),
            "Email": g(r, "Email"), "Phone Number": g(r, "Phone Number"),
            "LinkedIn Profile": g(r, "LinkedIn Profile"), "Company Domain": g(r, "Company Domain"),
        })
    if imported:
        cache_save(imported)   # upsert into SQLite by cache_key
    return len(imported)

def master_pull_from_drive():
    """Download the Drive master (if any) and merge it into the local SQLite. Safe no-op if Drive off."""
    if not drive_enabled():
        return 0
    try:
        token = _drive_access_token()
        fid = _drive_find_master(token)
        if not fid:
            return 0
        tmp = os.path.join(os.path.dirname(MASTER_DB), "_master_pull.xlsx")
        _drive_download(token, fid, tmp)
        n = _import_master_xlsx_into_db(tmp)
        try: os.remove(tmp)
        except Exception: pass
        return n
    except Exception as e:
        log(f"      [drive] master pull failed: {e}")
        return 0

def master_push_to_drive():
    """Export the whole SQLite cache to the single permanent master file in Drive (create or overwrite)."""
    if not drive_enabled():
        return None
    try:
        token = _drive_access_token()
        tmp = os.path.join(os.path.dirname(MASTER_DB), "_master_push.xlsx")
        export_master_xlsx(tmp)
        fid = _drive_find_master(token)
        with open(tmp, "rb") as f:
            data = f.read()
        try: os.remove(tmp)
        except Exception: pass
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if fid:
            # overwrite existing master (keeps same file & link)
            r = requests.patch(GDRIVE_UPDATE_URL.format(fid=fid),
                               headers={"Authorization": f"Bearer {token}", "Content-Type": mime},
                               data=data, timeout=120)
            r.raise_for_status()
        else:
            # create it once
            boundary = "----bimmaster7hf83n"
            meta = {"name": MASTER_DRIVE_NAME, "parents": [os.environ["GDRIVE_FOLDER_ID"]]}
            body = ((f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{json.dumps(meta)}\r\n").encode()
                    + f"--{boundary}\r\nContent-Type: {mime}\r\n\r\n".encode() + data
                    + f"\r\n--{boundary}--".encode())
            r = requests.post(GDRIVE_UPLOAD_URL,
                              headers={"Authorization": f"Bearer {token}",
                                       "Content-Type": f"multipart/related; boundary={boundary}"},
                              data=body, timeout=120)
            r.raise_for_status()
            fid = r.json().get("id")
        return f"https://drive.google.com/file/d/{fid}/view" if fid else None
    except Exception as e:
        log(f"      [drive] master push failed: {e}")
        return None

def sync_master_before_run():
    """Pull the permanent Drive master into local cache before a run (thread-safe)."""
    with _MASTER_SYNC_LOCK:
        return master_pull_from_drive()

def sync_master_after_run():
    """Push the updated local cache back to the permanent Drive master (thread-safe)."""
    with _MASTER_SYNC_LOCK:
        return master_push_to_drive()


# ---------------------------------------------------------------------------
# WATCHED FOLDER (hot folder): drop a CSV/XLSX in the Input folder -> it is
# processed automatically and results land in the Output folder.
#   Env: GDRIVE_INPUT_FOLDER, GDRIVE_OUTPUT_FOLDER, AUTO_MAX_COMPANIES (safety cap)
# A processed input is renamed with a "[done] " prefix so it is never re-run.
# ---------------------------------------------------------------------------
DONE_PREFIX = "[done] "
SKIP_PREFIX = "[skipped] "
AUTO_MAX_COMPANIES = int(os.environ.get("AUTO_MAX_COMPANIES", "50"))

def watch_enabled():
    return drive_enabled() and bool(os.environ.get("GDRIVE_INPUT_FOLDER")) and bool(os.environ.get("GDRIVE_OUTPUT_FOLDER"))

def _drive_list_inputs(token):
    """Return [{id,name,mimeType}] of un-processed CSV/XLSX files in the input folder."""
    folder = os.environ["GDRIVE_INPUT_FOLDER"]
    q = (f"'{folder}' in parents and trashed=false")
    r = requests.get(GDRIVE_FILES_URL, headers={"Authorization": f"Bearer {token}"},
                     params={"q": q, "fields": "files(id,name,mimeType)",
                             "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"},
                     timeout=30)
    r.raise_for_status()
    out = []
    for f in r.json().get("files", []):
        nm = f.get("name", "")
        low = nm.lower()
        if nm.startswith(DONE_PREFIX) or nm.startswith(SKIP_PREFIX):
            continue
        if low.startswith(UNIVERSAL_INPUT_BASE.lower()):
            continue  # the permanent universal input is handled by scan_universal_once
        if low.endswith(".csv") or low.endswith(".xlsx") or low.endswith(".xlsm") or low.endswith(".tsv"):
            out.append(f)
    return out

def _drive_rename(token, fid, new_name):
    r = requests.patch(f"{GDRIVE_FILES_URL}/{fid}",
                       headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                       params={"supportsAllDrives": "true"},
                       data=json.dumps({"name": new_name}), timeout=30)
    r.raise_for_status()

def _drive_upload_to(token, folder_id, local_path, drive_name, mime):
    with open(local_path, "rb") as f:
        data = f.read()
    boundary = "----bimwatch7hf83n"
    meta = {"name": drive_name, "parents": [folder_id]}
    body = ((f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{json.dumps(meta)}\r\n").encode()
            + f"--{boundary}\r\nContent-Type: {mime}\r\n\r\n".encode() + data
            + f"\r\n--{boundary}--".encode())
    r = requests.post(GDRIVE_UPLOAD_URL, headers={"Authorization": f"Bearer {token}",
                      "Content-Type": f"multipart/related; boundary={boundary}"}, data=body, timeout=120)
    r.raise_for_status()
    return r.json().get("id")

def process_input_file(fmeta, api_key, logfn=None, max_contacts=None):
    """Download one input file, run the pipeline, upload results to the output folder,
    and mark the input as done. Uses the permanent master (reuse = free). Returns a summary dict."""
    def say(m):
        (logfn or log)(m)
    token = _drive_access_token()
    fid, name = fmeta["id"], fmeta["name"]
    workdir = os.path.dirname(MASTER_DB)
    local_in = os.path.join(workdir, "_auto_in_" + re.sub(r"[^A-Za-z0-9._-]", "_", name))
    _drive_download(token, fid, local_in)

    headers = auth_headers(api_key)
    companies = load_companies(local_in)
    capped = False
    if len(companies) > AUTO_MAX_COMPANIES:
        companies = companies[:AUTO_MAX_COMPANIES]
        capped = True
    say(f"[auto] {name}: {len(companies)} companies"
        + (f" (CAPPED at {AUTO_MAX_COMPANIES})" if capped else ""))

    # pull permanent master so already-known companies are reused free
    sync_master_before_run()

    all_rows, found, cached_n, nomatch = [], 0, 0, 0
    for i, (company, domain) in enumerate(companies, 1):
        try:
            rows, kind = process_company(company, headers,
                                         max_contacts=max_contacts or MAX_CONTACTS_PER_COMPANY,
                                         company_domain=domain)
        except Exception as e:
            rows, kind = [note_row(company, f"ERROR: {e}")], "error"
        all_rows.extend(rows)
        if kind == "cached": cached_n += len([r for r in rows if r.get("First Name")]); found += 1
        elif kind == "found": found += 1
        elif kind == "nomatch" or kind == "nocontacts": nomatch += 1
        say(f"[auto] {name}: {i}/{len(companies)} {company}"
            + (f" [{domain}]" if domain else "")
            + (" (from master, free)" if kind == "cached" else ""))

    # write + upload results
    out_local = os.path.join(workdir, "_auto_out.xlsx")
    write_xlsx(all_rows, out_local)
    import datetime as _dt
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    base = re.sub(r"\.(csv|xlsx|xlsm|tsv)$", "", name, flags=re.I)
    out_name = f"RESULTS_{base}_{stamp}.xlsx"
    out_id = _drive_upload_to(token, os.environ["GDRIVE_OUTPUT_FOLDER"], out_local, out_name,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # update permanent master with anything newly researched
    sync_master_after_run()
    # mark input as done
    try: _drive_rename(token, fid, DONE_PREFIX + name)
    except Exception: pass
    for p in (local_in, out_local):
        try: os.remove(p)
        except Exception: pass

    say(f"[auto] {name}: DONE - {len(all_rows)} rows, {cached_n} reused free, saved '{out_name}' to Output.")
    return {"input": name, "output": out_name, "output_id": out_id, "rows": len(all_rows),
            "reused_free": cached_n, "capped": capped}

def scan_input_folder_once(api_key, logfn=None, max_contacts=None):
    """Process every new file currently in the input folder. Returns list of summaries."""
    if not watch_enabled():
        return []
    token = _drive_access_token()
    files = _drive_list_inputs(token)
    results = []
    for fmeta in files:
        try:
            results.append(process_input_file(fmeta, api_key, logfn=logfn, max_contacts=max_contacts))
        except Exception as e:
            (logfn or log)(f"[auto] {fmeta.get('name')}: ERROR {e}")
    return results


# ---------------------------------------------------------------------------
# UNIVERSAL INPUT/OUTPUT FILES: one permanent input file the team pastes
# companies into, one permanent output file that accumulates results.
#   - Input:  a file named "BIM_UNIVERSAL_INPUT..." in the Input folder
#             (auto-created as .xlsx with a "Company" header if missing).
#   - Quiet timer: only processed after the file has been UNCHANGED for
#     UNIVERSAL_QUIET_SECONDS (default 180s) so half-pasted lists aren't run.
#   - Only NEW companies (not in the output's Processed sheet) are processed.
#   - Output: "BIM_UNIVERSAL_OUTPUT.xlsx" in the Output folder, appended forever.
# ---------------------------------------------------------------------------
UNIVERSAL_INPUT_BASE = "BIM_UNIVERSAL_INPUT"
UNIVERSAL_OUTPUT_NAME = "BIM_UNIVERSAL_OUTPUT.xlsx"
UNIVERSAL_QUIET_SECONDS = int(os.environ.get("UNIVERSAL_QUIET_SECONDS", "180"))
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _drive_find_by_prefix(token, folder_id, prefix):
    q = f"'{folder_id}' in parents and trashed=false"
    r = requests.get(GDRIVE_FILES_URL, headers={"Authorization": f"Bearer {token}"},
                     params={"q": q, "fields": "files(id,name,mimeType,modifiedTime)",
                             "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"},
                     timeout=30)
    r.raise_for_status()
    for f in r.json().get("files", []):
        if f.get("name", "").lower().startswith(prefix.lower()):
            return f
    return None

def _parse_drive_time(s):
    # e.g. 2026-07-15T16:00:00.000Z
    from datetime import datetime, timezone
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0

def _universal_ensure_input(token, logfn=None):
    """Find the universal input file; create an empty template if missing."""
    f = _drive_find_by_prefix(token, os.environ["GDRIVE_INPUT_FOLDER"], UNIVERSAL_INPUT_BASE)
    if f:
        return f
    tmp = os.path.join(os.path.dirname(MASTER_DB), "_universal_tpl.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Companies"
    ws.append(["Company", "Website", "Email"])
    # Add a short instruction row so anyone opening the file understands the columns.
    # Loader skips rows whose Company cell is wrapped in parentheses.
    ws.append(["(paste company name here, one per row)",
               "(optional: exact website like bcsconstructiongroup.com)",
               "(optional: any email at the company like john@bcsconstructiongroup.com - fill EITHER Website OR Email, not both)"])
    wb.save(tmp)
    fid = _drive_upload_to(token, os.environ["GDRIVE_INPUT_FOLDER"], tmp,
                           UNIVERSAL_INPUT_BASE + ".xlsx", XLSX_MIME)
    try: os.remove(tmp)
    except Exception: pass
    (logfn or log)(f"[universal] Created '{UNIVERSAL_INPUT_BASE}.xlsx' in the Input folder - "
                   "open it, put company names in 'Company', and (optionally) fill EITHER "
                   "'Website' (e.g. bcsconstructiongroup.com) OR 'Email' (e.g. john@bcs...) "
                   "for stricter, correct-company matching.")
    return {"id": fid, "name": UNIVERSAL_INPUT_BASE + ".xlsx", "mimeType": XLSX_MIME,
            "modifiedTime": "1970-01-01T00:00:00Z"}

def _universal_read_companies(token, f):
    """Read company names from the universal input (Google Sheet or Excel/CSV)."""
    workdir = os.path.dirname(MASTER_DB)
    if f.get("mimeType") == GSHEET_MIME:
        # native Google Sheet -> export as CSV
        r = requests.get(f"{GDRIVE_FILES_URL}/{f['id']}/export",
                         headers={"Authorization": f"Bearer {token}"},
                         params={"mimeType": "text/csv"}, timeout=60)
        r.raise_for_status()
        tmp = os.path.join(workdir, "_universal_in.csv")
        with open(tmp, "wb") as fh: fh.write(r.content)
    else:
        ext = ".csv" if f.get("name", "").lower().endswith(".csv") else ".xlsx"
        tmp = os.path.join(workdir, "_universal_in" + ext)
        _drive_download(token, f["id"], tmp)
    try:
        return load_companies(tmp)
    finally:
        try: os.remove(tmp)
        except Exception: pass

def _universal_load_output(token):
    """Return (fileId or None, existing rows, processed-company set) from the universal output."""
    f = _drive_find_by_prefix(token, os.environ["GDRIVE_OUTPUT_FOLDER"], UNIVERSAL_OUTPUT_NAME)
    if not f:
        return None, [], set()
    workdir = os.path.dirname(MASTER_DB)
    tmp = os.path.join(workdir, "_universal_out.xlsx")
    _drive_download(token, f["id"], tmp)
    rows, processed = [], set()
    try:
        wb = load_workbook(tmp, read_only=True, data_only=True)
        ws = wb["Contacts"] if "Contacts" in wb.sheetnames else wb.active
        data = list(ws.iter_rows(values_only=True))
        if data:
            hdr = [str(h) if h else "" for h in data[0]]
            for r in data[1:]:
                if any(v not in (None, "") for v in r):
                    rows.append({hdr[i]: (r[i] if i < len(r) else "") for i in range(len(hdr))})
        if "Processed" in wb.sheetnames:
            for r in wb["Processed"].iter_rows(values_only=True):
                if r and r[0]:
                    processed.add(_norm(str(r[0])))
    finally:
        try: os.remove(tmp)
        except Exception: pass
    return f["id"], rows, processed

def _universal_write_output(token, fid, rows, processed):
    """Write the universal output workbook (Contacts + Processed sheets) to Drive."""
    workdir = os.path.dirname(MASTER_DB)
    tmp = os.path.join(workdir, "_universal_out_w.xlsx")
    wb = Workbook()
    ws = wb.active; ws.title = "Contacts"; ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row.get(k, "") for k in OUTPUT_COLUMNS])
    ps = wb.create_sheet("Processed")
    for name in sorted(processed):
        ps.append([name])
    wb.save(tmp)
    try:
        if fid:
            r = requests.patch(GDRIVE_UPDATE_URL.format(fid=fid),
                               headers={"Authorization": f"Bearer {token}", "Content-Type": XLSX_MIME},
                               data=open(tmp, "rb").read(), timeout=120)
            r.raise_for_status()
        else:
            fid = _drive_upload_to(token, os.environ["GDRIVE_OUTPUT_FOLDER"], tmp,
                                   UNIVERSAL_OUTPUT_NAME, XLSX_MIME)
    finally:
        try: os.remove(tmp)
        except Exception: pass
    return fid

_UNIVERSAL_STATE = {"input_mtime": "", "processed": None, "checked_at": 0.0,
                    "last_status_at": 0.0}

def scan_universal_once(api_key, logfn=None, max_contacts=None):
    """Check the universal input file. If it changed and has been quiet long enough,
    process only the NEW companies and append them to the universal output.

    Fast-path: if the input's Drive modifiedTime hasn't changed since our last check
    AND we have a cached 'processed' set, we skip the expensive download entirely.
    A status line is logged at most once every ~5 min so 'nothing happening' scans
    still leave visible traces without spamming the log."""
    def say(m): (logfn or log)(m)
    def status(m):
        # heartbeat: prints once every 5 min so /watch-status always shows recent state
        if time.time() - _UNIVERSAL_STATE["last_status_at"] > 300:
            say(m); _UNIVERSAL_STATE["last_status_at"] = time.time()
    if not watch_enabled():
        return []
    try:
        token = _drive_access_token()
        f = _universal_ensure_input(token, logfn=logfn)
    except Exception as e:
        say(f"[universal] could not reach Drive: {e}")
        return []

    input_mtime = f.get("modifiedTime", "")
    age = time.time() - _parse_drive_time(input_mtime)

    # QUIET TIMER: only skip if the file was edited too recently (someone may still be typing)
    if age < UNIVERSAL_QUIET_SECONDS:
        status(f"[universal] input edited {int(age)}s ago - waiting for it to be quiet "
               f"({UNIVERSAL_QUIET_SECONDS}s) before processing.")
        return []

    # FAST PATH: input hasn't changed since our last successful scan -> nothing to do,
    # skip the whole Drive download/upload round trip.
    if (_UNIVERSAL_STATE["input_mtime"] == input_mtime
            and _UNIVERSAL_STATE["processed"] is not None):
        status(f"[universal] input unchanged (mtime {input_mtime}) - nothing to do.")
        return []

    say(f"[universal] input changed (edited {int(age)}s ago) - reading it now.")
    try:
        companies = _universal_read_companies(token, f)
    except Exception as e:
        say(f"[universal] could not read input file: {e}")
        return []
    say(f"[universal] input has {len(companies)} company name(s) total.")

    fid, rows, processed = _universal_load_output(token)
    # companies is now a list of (name, domain) tuples
    new = [(c, d) for (c, d) in companies if _norm(c) not in processed]
    if not new:
        # remember state so future scans can skip the download until input changes again
        _UNIVERSAL_STATE.update({"input_mtime": input_mtime, "processed": processed,
                                 "checked_at": time.time()})
        return []

    # CHUNKING for the free tier's memory ceiling: process just a few companies per
    # scan cycle, save immediately, then let the NEXT scan pick up where we left off.
    # A larger batch would try to hold too much in RAM at once and get the worker
    # killed. Set UNIVERSAL_CHUNK_SIZE higher (e.g. 5-10) on paid tiers.
    chunk_size = int(os.environ.get("UNIVERSAL_CHUNK_SIZE", "2"))
    total_pending = len(new)
    capped_by_safety = len(new) > AUTO_MAX_COMPANIES
    new = new[:AUTO_MAX_COMPANIES]
    this_batch = new[:chunk_size]
    remaining_after = len(new) - len(this_batch)
    say(f"[universal] {total_pending} new company(ies) pending"
        + (" (CAPPED)" if capped_by_safety else "")
        + f". Processing {len(this_batch)} this cycle; "
        + (f"{remaining_after} more will run on the next scan(s)." if remaining_after
           else "this is the last chunk."))

    sync_master_before_run()
    headers = auth_headers(api_key)
    done = 0
    for company, domain in this_batch:
        try:
            crows, kind = process_company(company, headers,
                                          max_contacts=max_contacts or MAX_CONTACTS_PER_COMPANY,
                                          company_domain=domain)
        except Exception as e:
            crows, kind = [note_row(company, f"ERROR: {e}")], "error"
        rows.extend(crows)
        processed.add(_norm(company))
        done += 1
        say(f"[universal] {done}/{len(this_batch)} {company}"
            + (f" [{domain}]" if domain else "")
            + (" (from master, free)" if kind == "cached" else ""))

    fid = _universal_write_output(token, fid, rows, processed)
    sync_master_after_run()
    # Don't mark input_mtime as fully "handled" while there are still companies pending
    # - that way the NEXT scan cycle picks them up automatically.
    if remaining_after:
        say(f"[universal] chunk done - {remaining_after} company(ies) still to go, "
            f"they will process on the next scan.")
        # leave input_mtime unchanged in state so we re-enter the loop next scan
        _UNIVERSAL_STATE.update({"processed": processed, "checked_at": time.time()})
    else:
        _UNIVERSAL_STATE.update({"input_mtime": input_mtime, "processed": processed,
                                 "checked_at": time.time()})
        say(f"[universal] ALL DONE - {len(rows)} contact rows in {UNIVERSAL_OUTPUT_NAME}.")
    return [{"new_companies": done, "output": UNIVERSAL_OUTPUT_NAME,
             "capped": capped_by_safety, "remaining": remaining_after}]
    # Update fast-path state to the current mtime so we don't re-scan until input changes
    _UNIVERSAL_STATE.update({"input_mtime": input_mtime, "processed": processed,
                             "checked_at": time.time()})
    say(f"[universal] DONE - {done} company(ies) added to {UNIVERSAL_OUTPUT_NAME}.")
    return [{"new_companies": done, "output": UNIVERSAL_OUTPUT_NAME, "capped": capped}]


def diagnose():
    """Plain-English check of the watched-folder setup. Returns a dict of findings."""
    out = {"checks": []}
    def add(ok, msg): out["checks"].append(("OK" if ok else "PROBLEM") + ": " + msg)
    if not drive_enabled():
        add(False, "Google Drive credentials are not all set on the server.")
        return out
    add(True, "Drive credentials are present.")
    try:
        token = _drive_access_token()
        add(True, "Drive login (refresh token) works.")
    except Exception as e:
        add(False, f"Drive login FAILED: {e}. The refresh token may be expired - re-run get_drive_token.py.")
        return out
    if not (os.environ.get("GDRIVE_INPUT_FOLDER") and os.environ.get("GDRIVE_OUTPUT_FOLDER")):
        add(False, "Input/Output folder IDs are not set on the server.")
        return out
    add(True, "Input/Output folder IDs are set.")
    # what does the automation account see in the input folder?
    try:
        folder = os.environ["GDRIVE_INPUT_FOLDER"]
        r = requests.get(GDRIVE_FILES_URL, headers={"Authorization": f"Bearer {token}"},
                         params={"q": f"'{folder}' in parents and trashed=false",
                                 "fields": "files(id,name,mimeType,owners(emailAddress))",
                                 "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"},
                         timeout=30)
        r.raise_for_status()
        files = r.json().get("files", [])
        out["input_folder_id"] = folder
        out["files_visible_to_app"] = [
            {"name": f.get("name"), "owner": (f.get("owners", [{}])[0].get("emailAddress", "?"))}
            for f in files]
        if files:
            add(True, f"App can SEE {len(files)} file(s) in the Input folder.")
            proc = [f for f in files if not f["name"].startswith(DONE_PREFIX)]
            add(True, f"{len(proc)} of them are unprocessed and would be picked up.")
        else:
            add(False, "App sees ZERO files in the Input folder. Likely cause: the file was "
                       "uploaded by a DIFFERENT Google account, so this app (drive.file scope) "
                       "can't see it. Fix: broaden the scope, or have files created by the app's account.")
    except Exception as e:
        add(False, f"Could not list the Input folder: {e}")
    # can we write to output?
    try:
        tmp = os.path.join(os.path.dirname(MASTER_DB), "_diag.xlsx")
        wb = Workbook(); wb.active.append(["diag"]); wb.save(tmp)
        fid = _drive_upload_to(token, os.environ["GDRIVE_OUTPUT_FOLDER"], tmp, "_connection_test.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        try: os.remove(tmp)
        except Exception: pass
        add(bool(fid), "Wrote a test file to the Output folder (delete '_connection_test.xlsx' later).")
    except Exception as e:
        add(False, f"Could NOT write to the Output folder: {e}")
    return out


def main():
    ap = argparse.ArgumentParser(description="BIM Data Execution - all hierarchy contacts per company")
    ap.add_argument("--input", required=True); ap.add_argument("--output", default="results.xlsx")
    ap.add_argument("--api-key", default=os.environ.get("SEAMLESS_API_KEY", ""))
    ap.add_argument("--max-contacts", type=int, default=MAX_CONTACTS_PER_COMPANY)
    args = ap.parse_args()
    if not args.api_key:
        log("ERROR: no API key. Pass --api-key or set SEAMLESS_API_KEY."); sys.exit(1)
    headers = auth_headers(args.api_key)
    companies = load_companies(args.input)
    log(f"Loaded {len(companies)} companies.")
    all_rows = []
    for i, (company, domain) in enumerate(companies, 1):
        log(f"[{i}/{len(companies)}] {company}" + (f" ({domain})" if domain else ""))
        try:
            rows, kind = process_company(company, headers, max_contacts=args.max_contacts,
                                         company_domain=domain)
        except Exception as e:
            rows = [note_row(company, f"ERROR: {e}")]
        all_rows.extend(rows)
        for r in rows:
            who = f"{r['First Name']} {r['Last Name']}".strip()
            log(f"      {r.get('Hierarchy Role') or ''}: {who or r['Note']}")
        try: write_xlsx(all_rows, args.output)
        except Exception: pass
        time.sleep(DELAY_BETWEEN_COMPANIES)
    write_xlsx(all_rows, args.output)
    log(f"\nDone. Wrote {len(all_rows)} rows to {args.output}")


if __name__ == "__main__":
    main()